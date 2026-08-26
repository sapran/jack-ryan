"""Extractors for tabular files.

Rendered so that a passage returned to a reader shows which sheet and which row
it came from. An undifferentiated run of values is worse than useless in a
citation: it names no place a person could look.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from .extractors import Extraction, ExtractionError

# A row past this many cells is truncated. A spreadsheet with ten thousand
# columns is a generated artefact, and rendering it whole would bury the sheet.
MAX_CELLS_PER_ROW = 512


def _render_row(number: int, values: list[str]) -> str:
    cells = [" ".join(str(v).split()) for v in values[:MAX_CELLS_PER_ROW] if v is not None]
    if not any(cells):
        return ""
    return f"row {number}: " + " | ".join(cells)


class SpreadsheetExtractor:
    """XLSX workbooks, sheet by sheet."""

    name = "spreadsheet"
    suffixes = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
    }

    def accepts(self, path: Path) -> bool:
        return path.suffix.lower() in self.suffixes

    def extract(self, path: Path) -> Extraction:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - packaging failure
            raise ExtractionError("the openpyxl reader is not installed") from exc
        try:
            # read_only streams rather than building the whole sheet in memory;
            # data_only takes cached values rather than formula source.
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise ExtractionError(
                f"could not read {path.name} as a workbook: {type(exc).__name__}: {exc}"
            ) from exc

        blocks: list[str] = []
        try:
            for sheet in workbook.worksheets:
                lines = [f"## sheet: {' '.join(str(sheet.title).split())}"]
                for number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    rendered = _render_row(number, list(row))
                    if rendered:
                        lines.append(rendered)
                if len(lines) > 1:
                    blocks.append("\n".join(lines))
                else:
                    # An empty sheet is named and left at that, so a workbook
                    # with one empty sheet still ingests on the strength of the
                    # others.
                    blocks.append(f"{lines[0]}\n(empty)")
        finally:
            workbook.close()

        return Extraction(
            text="\n\n".join(blocks),
            media_type=self.suffixes[path.suffix.lower()],
            extractor=self.name,
            metadata={"sheets": str(len(blocks))},
        )


class DelimitedExtractor:
    """CSV and TSV, rendered on the same shape as a workbook sheet."""

    name = "delimited"
    suffixes = {".csv": "text/csv", ".tsv": "text/tab-separated-values"}

    def accepts(self, path: Path) -> bool:
        return path.suffix.lower() in self.suffixes

    def extract(self, path: Path) -> Extraction:
        suffix = path.suffix.lower()
        delimiter = "\t" if suffix == ".tsv" else ","
        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            raise ExtractionError(f"could not read {path.name}: {exc}") from exc

        lines = [f"## sheet: {' '.join(path.stem.split())}"]
        reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
        try:
            for number, row in enumerate(reader, start=1):
                rendered = _render_row(number, list(row))
                if rendered:
                    lines.append(rendered)
        except csv.Error as exc:
            raise ExtractionError(f"could not parse {path.name}: {exc}") from exc

        return Extraction(
            text="\n".join(lines),
            media_type=self.suffixes[suffix],
            extractor=self.name,
        )
